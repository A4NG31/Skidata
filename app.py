import pandas as pd
import streamlit as st
import io
from openpyxl import load_workbook
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime

# -------------------------
# Configuración página
# -------------------------
st.set_page_config(
    page_title="GoPass - Validador Terpel",
    page_icon="⛽",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ================================
# CSS personalizado GoPass
# ================================
st.markdown("""
<style>
    :root {
        --gopass-blue: #1E3A8A;
        --gopass-light-blue: #3B82F6;
        --gopass-orange: #F59E0B;
        --gopass-green: #10B981;
        --gopass-red: #EF4444;
        --gopass-gray: #6B7280;
        --gopass-light-gray: #F3F4F6;
    }
    header[data-testid="stHeader"], .stDeployButton {display: none;}
    .stMainBlockContainer {padding-top: 1rem;}

    .main-header {
        background: linear-gradient(135deg, var(--gopass-blue) 0%, var(--gopass-light-blue) 100%);
        color: white;
        padding: 2rem 0;
        margin: -1rem -1rem 2rem -1rem;
        text-align: center;
        border-radius: 0 0 20px 20px;
        box-shadow: 0 4px 20px rgba(30, 58, 138, 0.3);
    }
    .main-title { font-size: 2.5rem; font-weight: 700; margin-bottom: 0.5rem; text-shadow: 2px 2px 4px rgba(0,0,0,0.3); }
    .main-subtitle { font-size: 1.2rem; opacity: 0.9; font-weight: 300; }

    .metric-card { background: white; padding: 1.5rem; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border-left: 4px solid var(--gopass-light-blue); margin-bottom: 1rem; }
    .custom-info { background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(59, 130, 246, 0.05)); border-left: 4px solid var(--gopass-light-blue); padding: 1rem 1.5rem; border-radius: 10px; margin: 1rem 0; color: var(--gopass-blue); }
    .custom-success { background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(16, 185, 129, 0.05)); border-left: 4px solid var(--gopass-green); padding: 1rem 1.5rem; border-radius: 10px; margin: 1rem 0; color: var(--gopass-green); }
    .custom-warning { background: linear-gradient(135deg, rgba(245, 158, 11, 0.1), rgba(245, 158, 11, 0.05)); border-left: 4px solid var(--gopass-orange); padding: 1rem 1.5rem; border-radius: 10px; margin: 1rem 0; color: var(--gopass-orange); }

    .footer { text-align: center; padding: 2rem 0; margin-top: 3rem; border-top: 1px solid var(--gopass-light-gray); color: var(--gopass-gray); font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)

# ================================
# Header
# ================================
st.markdown("""
<div class="main-header">
    <img src="https://i.imgur.com/z9xt46F.jpeg" 
         style="width: 200px; border-radius: 15px; margin-bottom: 1rem; box-shadow: 0 4px 15px rgba(0,0,0,0.3);" 
         alt="Logo GoPass">
    <div class="main-title">⛽ Validador de Dobles Cobros</div>
    <div class="main-subtitle">Gasolineras Terpel - Sistema de Detección Avanzada</div>
</div>
""", unsafe_allow_html=True)

# -------------------------
# Session state inicial
# -------------------------
if 'datos_procesados' not in st.session_state:
    st.session_state.datos_procesados = False
    st.session_state.df_final = None
    st.session_state.current_file = None
    st.session_state.stats = None

# ================================
# Upload UI
# ================================
st.markdown('<div style="margin-top:1rem;"></div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.markdown("### 📂 Cargar Base de Datos de Transacciones")
    st.markdown('<div class="custom-info">📋 Formatos soportados: Excel (.xlsx) | Tamaño máximo recomendado: 500MB</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Selecciona el archivo de transacciones Terpel", type=['xlsx'], accept_multiple_files=False)

# Reiniciar estado si cambia el archivo
if uploaded_file is not None and st.session_state.current_file != uploaded_file.name:
    st.session_state.datos_procesados = False
    st.session_state.df_final = None
    st.session_state.current_file = uploaded_file.name
    st.session_state.stats = None

# ================================
# Procesamiento del archivo (con progreso)
# ================================
if uploaded_file is not None and not st.session_state.datos_procesados:
    try:
        # verificar tamaño (en MB)
        file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
        if file_size_mb > 500:
            st.markdown(f'<div class="custom-warning">⚠️ Archivo muy grande ({file_size_mb:.1f} MB). Considera archivos más pequeños.</div>', unsafe_allow_html=True)
            st.stop()

        st.markdown(f'<div class="custom-success">📊 Archivo cargado: {uploaded_file.name} ({file_size_mb:.1f} MB)</div>', unsafe_allow_html=True)

        # show progress
        st.markdown("### ⚙️ Procesando Datos")
        progress_bar = st.progress(0)
        status_text = st.empty()

        # cargar con openpyxl para mostrar progreso row-by-row
        wb = load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        total_filas = ws.max_row
        total_registros_bruto = max(0, total_filas - 1)

        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            data.append(row)
            if i % 500 == 0 or i == total_filas:
                progress = int((i / total_filas) * 100)
                progress_bar.progress(progress)
                status_text.text(f"Lectura fila {i:,} de {total_filas:,}")

        status_text.text("✅ Archivo leído, creando DataFrame")
        columnas = data[0]
        df = pd.DataFrame(data[1:], columns=columnas)

        # Guardar total bruto (actual)
        total_registros_bruto = len(df)

        # Normalizar columnas críticas a tipos apropiados
        # Intentamos convertir montos a numérico (si vienen como texto)
        if 'Valor Pagado' in df.columns:
            df['Valor Pagado'] = pd.to_numeric(df['Valor Pagado'], errors='coerce')
        if 'Valor Servicio' in df.columns:
            df['Valor Servicio'] = pd.to_numeric(df['Valor Servicio'], errors='coerce')
        # Fecha de Pago
        if 'Fecha de Pago' in df.columns:
            df['Fecha de Pago'] = pd.to_datetime(df['Fecha de Pago'], dayfirst=True, errors='coerce')

        # Filtrado base (mismo criterio que en tu código original)
        df = df[(df.get('Valor Pagado', 0) > 0) & (df.get('Estado', '') == 'Exitosa')].copy()

        # Garantizar columnas necesarias existen
        needed = ['Fecha de Pago', 'Id', 'Establecimiento', 'Placa', 'Valor Servicio', 'Valor Pagado', 'Estado']
        for c in needed:
            if c not in df.columns:
                df[c] = pd.NA

        # Inicializar Novedad
        df['Novedad'] = "NORMAL"

        # Orden y reset
        df.sort_values(by=['Establecimiento', 'Placa', 'Valor Pagado', 'Fecha de Pago'], inplace=True, na_position='last')
        df.reset_index(drop=True, inplace=True)

        # Detección eficiente (recorriendo filas como en tu versión)
        total = len(df)
        dobles_detectados = 0
        if total > 1:
            for i in range(1, total):
                # Evitar errores si hay NaT o NaN
                try:
                    mismo_establecimiento = df.at[i, 'Establecimiento'] == df.at[i-1, 'Establecimiento']
                    misma_placa = df.at[i, 'Placa'] == df.at[i-1, 'Placa']
                    mismo_valor_servicio = df.at[i, 'Valor Servicio'] == df.at[i-1, 'Valor Servicio']
                    mismo_valor_pagado = df.at[i, 'Valor Pagado'] == df.at[i-1, 'Valor Pagado']
                    id_diferente = str(df.at[i, 'Id']) != str(df.at[i-1, 'Id'])
                    fecha_i = df.at[i, 'Fecha de Pago']
                    fecha_prev = df.at[i-1, 'Fecha de Pago']
                    if pd.isna(fecha_i) or pd.isna(fecha_prev):
                        dentro_tolerancia = False
                    else:
                        diferencia_tiempo = abs((fecha_i - fecha_prev).total_seconds())
                        dentro_tolerancia = diferencia_tiempo <= 600  # 10 minutos
                except Exception:
                    # en caso de cualquier problema comparativo, marcar como no doble
                    dentro_tolerancia = False
                    mismo_establecimiento = misma_placa = mismo_valor_servicio = mismo_valor_pagado = id_diferente = False

                if (mismo_establecimiento and misma_placa and mismo_valor_servicio and
                    mismo_valor_pagado and id_diferente and dentro_tolerancia):
                    df.at[i, 'Novedad'] = "DOBLE COBRO"
                    df.at[i-1, 'Novedad'] = "DOBLE COBRO"
                    dobles_detectados += 2

                # Actualizar progreso de análisis
                if i % 500 == 0 or i == total - 1:
                    progress = int((i / total) * 100)
                    progress_bar.progress(progress)
                    status_text.text(f"Analizando registro {i:,} de {total:,} - Dobles detectados: {dobles_detectados}")

        status_text.text("✅ Análisis completado")
        progress_bar.progress(100)

        # Preparar DataFrame final y estadísticas
        columnas_finales = ['Fecha de Pago', 'Id', 'Establecimiento', 'Placa', 'Valor Servicio',
                           'Valor Pagado', 'Estado', 'Novedad']
        st.session_state.df_final = df[columnas_finales].copy()

        total_registros = len(st.session_state.df_final)
        dobles_cobros = len(st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO'])
        porcentaje_dobles = (dobles_cobros / total_registros * 100) if total_registros > 0 else 0
        valor_total_dobles = st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO']['Valor Pagado'].sum()

        st.session_state.stats = {
            'total_registros_bruto': total_registros_bruto,
            'total_registros': total_registros,
            'dobles_cobros': dobles_cobros,
            'porcentaje_dobles': porcentaje_dobles,
            'valor_total_dobles': valor_total_dobles,
            'registros_normales': total_registros - dobles_cobros
        }
        st.session_state.datos_procesados = True

    except Exception as e:
        st.markdown(f'<div class="custom-warning">⚠️ Error al procesar el archivo: {str(e)}</div>', unsafe_allow_html=True)
        st.stop()

# ================================
# Mostrar resultados (dashboard + gráficas + descarga)
# ================================
if uploaded_file is not None and st.session_state.datos_procesados and st.session_state.df_final is not None:
    st.markdown("---")
    st.markdown("### 📊 Dashboard de Resultados")

    stats = st.session_state.get("stats", {})

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("TOTAL REGISTROS", f"{stats.get('total_registros_bruto', 0):,}".replace(",", "."))
    col2.metric("TOTAL FILTRADOS", f"{stats.get('total_registros', 0):,}".replace(",", "."))
    col3.metric("DOBLES COBROS", f"{stats.get('dobles_cobros', 0):,}".replace(",", "."))
    col4.metric("% DOBLES COBROS", f"{stats.get('porcentaje_dobles', 0):.2f}%")
    col5.metric("VALOR TOTAL DOBLES", f"${stats.get('valor_total_dobles', 0):,}".replace(",", "."))

    # Gráficas: Dona + Top establecimientos
    colA, colB = st.columns(2)
    with colA:
        fig_dona = go.Figure(data=[go.Pie(
            labels=['Registros Normales', 'Dobles Cobros'],
            values=[stats.get('registros_normales', 0), stats.get('dobles_cobros', 0)],
            hole=0.6,
            marker_colors=['#10B981', '#EF4444']
        )])
        fig_dona.update_layout(title="Distribución de Registros", title_x=0.5, height=350)
        st.plotly_chart(fig_dona, use_container_width=True)

    with colB:
        if stats.get('dobles_cobros', 0) > 0:
            top_establecimientos = st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO'] \
                                        ['Establecimiento'].value_counts().head(10)
            fig_bar = px.bar(
                x=top_establecimientos.values,
                y=top_establecimientos.index,
                orientation='h',
                title="Top 10 Establecimientos con Dobles Cobros",
                color=top_establecimientos.values,
                color_continuous_scale=['#F59E0B', '#EF4444']
            )
            fig_bar.update_layout(height=350, showlegend=False)
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("No hay dobles cobros para graficar.")

    # Vista previa y filtros
    st.markdown("### 📋 Vista Previa de Resultados")
    colF1, colF2, colF3 = st.columns(3)
    with colF1:
        filtro_novedad = st.selectbox("Filtrar por novedad:", ["Todos", "NORMAL", "DOBLE COBRO"])
    with colF2:
        num_registros = st.selectbox("Registros a mostrar:", [50, 100, 200, 500, "Todos"])
    df_mostrar = st.session_state.df_final.copy()
    if filtro_novedad != "Todos":
        df_mostrar = df_mostrar[df_mostrar['Novedad'] == filtro_novedad]
    if num_registros != "Todos":
        df_mostrar = df_mostrar.head(num_registros)

    st.dataframe(df_mostrar, use_container_width=True, height=420)

    # Descarga Excel con resumen y hojas
    colD1, colD2, colD3 = st.columns([1, 1, 1])
    with colD2:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            st.session_state.df_final.to_excel(writer, sheet_name='Resultados_Completos', index=False)
            dobles_df = st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO']
            dobles_df.to_excel(writer, sheet_name='Solo_Dobles_Cobros', index=False)
            resumen_data = {
                'Métrica': ['Total Registros', 'Dobles Cobros', 'Registros Normales', '% Dobles Cobros', 'Valor Total Dobles'],
                'Valor': [stats.get('total_registros', 0), stats.get('dobles_cobros', 0), stats.get('registros_normales', 0),
                          f"{stats.get('porcentaje_dobles', 0):.2f}%", f"${stats.get('valor_total_dobles', 0):,.0f}"]
            }
            pd.DataFrame(resumen_data).to_excel(writer, sheet_name='Resumen', index=False)
        buffer.seek(0)
        st.download_button(
            label="💾 Descargar Análisis Completo",
            data=buffer,
            file_name=f"Terpel_DoblesCobros_Analisis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Botón reiniciar
    if st.button("🔄 Analizar Nuevo Archivo"):
        st.session_state.datos_procesados = False
        st.session_state.df_final = None
        st.session_state.current_file = None
        st.session_state.stats = None
        st.experimental_rerun()

# Instrucciones iniciales si no hay archivo
if uploaded_file is None:
    st.markdown("""
    <div class="custom-info">
        <h4>📋 Instrucciones de Uso</h4>
        <p><strong>1.</strong> Descarga la base de transacciones Terpel desde la consola GoPass</p>
        <p><strong>2.</strong> Carga el archivo Excel (.xlsx) utilizando el botón superior</p>
        <p><strong>3.</strong> El sistema detectará automáticamente los dobles cobros basado en:</p>
        <ul>
            <li>✅ Mismo Establecimiento, Placa, Valor Servicio y Valor Pagado</li>
            <li>⏰ Diferencia de tiempo ≤ 10 minutos</li>
            <li>✅ Estado = "Exitosa"</li>
            <li>🆔 ID Transaction diferentes</li>
        </ul>
        <p><strong>4.</strong> Revisa los resultados en el dashboard y descarga el análisis</p>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div class="footer">
    <hr style="border: 1px solid var(--gopass-light-gray); margin: 1rem 0;">
    <p><strong>🚀 GoPass Analytics Platform</strong> | Validador de Dobles Cobros Terpel v2.0</p>
    <p>Desarrollado por <strong>Angel Torres</strong> | © 2025 GoPass</p>
</div>
""", unsafe_allow_html=True)
